from __future__ import annotations

import io
import re
from datetime import date
from typing import Any, Literal
from urllib.parse import urlencode

from fastapi import Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from backend import accountant_excel_api as excel
from backend import accountant_selected_plan_api as selected
from backend.accountant_drive import AccountantDriveUploadError, upload_accountant_artifact
from backend.db import connection

_INSTALLED = False


def _qty_format(value: Any) -> str:
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        return "General"
    return "#,##0" if abs(number - round(number)) < 0.0000001 else "#,##0.####"


def polished_workbook_bytes(rows: list[list[Any]]) -> bytes:
    """Create a clean accountant workbook with stable numeric formatting.

    Integer quantities no longer render with a trailing decimal separator, while
    fractional quantities remain numeric. Excel localizes group/decimal separators
    according to the user's regional settings.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Belanja"

    for row in rows:
        ws.append(row)

    blue = "1E40AF"
    blue_dark = "1E3A8A"
    blue_soft = "DBEAFE"
    zebra = "F8FAFC"
    border_color = "CBD5E1"
    thin = Side(style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 34

    summary_labels = {"GRAND TOTAL", "PAGU BGN", "SELISIH PAGU - ESTIMASI"}
    last_data_row = 1
    for row_index in range(2, ws.max_row + 1):
        label = str(ws.cell(row_index, 1).value or "").upper()
        values = [ws.cell(row_index, column).value for column in range(1, 8)]
        if not any(value not in (None, "") for value in values):
            ws.row_dimensions[row_index].height = 8
            continue
        is_summary = label in summary_labels
        if not is_summary:
            last_data_row = row_index

        for column in range(1, 8):
            cell = ws.cell(row_index, column)
            cell.border = border
            if is_summary:
                cell.fill = PatternFill("solid", fgColor=blue_soft)
                cell.font = Font(bold=True, color=blue_dark)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                if row_index % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=zebra)
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        if is_summary:
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=4)
            ws.cell(row_index, 1).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row_index, 5).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row_index, 5).number_format = "#,##0"
            ws.row_dimensions[row_index].height = 23
        else:
            ws.cell(row_index, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.cell(row_index, 2).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row_index, 3).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row_index, 4).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row_index, 5).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row_index, 6).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.cell(row_index, 7).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.cell(row_index, 2).number_format = _qty_format(ws.cell(row_index, 2).value)
            ws.cell(row_index, 4).number_format = "#,##0"
            ws.cell(row_index, 5).number_format = "#,##0"
            ws.row_dimensions[row_index].height = 24

    widths = [34, 14, 13, 23, 24, 40, 25]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    if last_data_row >= 2:
        ws.auto_filter.ref = f"A1:G{last_data_row}"

    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _safe_excel_filename(value: str | None, default: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return default
    name = raw.replace("\\", "/").split("/")[-1]
    name = re.sub(r"[<>:\"/\\|?*\x00-\x1f]+", "_", name).strip(" .")
    if not name:
        return default
    if not name.lower().endswith(".xlsx"):
        name += ".xlsx"
    return name[:160]


class NamedSelectedPlanExcelIn(BaseModel):
    site: Literal["MAJA", "CEMPLANG"]
    distribution_date: date
    calculator_document_id: str = Field(min_length=1)
    custom_filename: str | None = Field(default=None, max_length=160)
    commit: bool = False


def _artifact(payload: NamedSelectedPlanExcelIn) -> dict[str, Any]:
    artifact = selected._selected_plan_artifact(
        payload.site,
        payload.distribution_date,
        payload.calculator_document_id,
    )
    artifact["filename"] = _safe_excel_filename(payload.custom_filename, artifact["filename"])
    return artifact


def _download_url(payload: NamedSelectedPlanExcelIn) -> str:
    params = {
        "site": payload.site,
        "distributionDate": payload.distribution_date.isoformat(),
        "calculatorDocumentId": payload.calculator_document_id,
    }
    if payload.custom_filename:
        params["customFilename"] = _safe_excel_filename(payload.custom_filename, "daftar_belanja.xlsx")
    return "/v1/accountant-excel/download-selected-plan?" + urlencode(params)


def named_selected_plan_excel(payload: NamedSelectedPlanExcelIn) -> dict[str, Any]:
    selected.require_db()
    artifact = _artifact(payload)
    accountant = excel.ACCOUNTANTS[payload.site]
    base = {
        "committed": False,
        "duplicate": False,
        "site": payload.site,
        "accountantCode": accountant,
        "distributionDate": payload.distribution_date.isoformat(),
        "calculatorDocumentId": payload.calculator_document_id,
        "planName": artifact["planName"],
        "planningSource": "CALCULATOR_DAILY_PLAN_SINGLE",
        "filename": artifact["filename"],
        "sheetName": "Belanja",
        "itemCount": len(artifact["items"]),
        "grandTotal": artifact["grandTotal"],
        "paguBgn": artifact["paguBgn"],
        "paguMinusEstimate": artifact["paguMinusEstimate"],
        "excelReady": True,
        "downloadAvailable": True,
        "downloadUrl": _download_url(payload),
        "fileSizeBytes": len(artifact["xlsx"]),
        "driveUri": None,
        "driveUploadStatus": "NOT_REQUESTED" if not payload.commit else "PENDING",
        "driveUploadError": None,
        "retryable": False,
        "excelFormatVersion": "accountant-polished-v1",
    }
    if not payload.commit:
        return base

    with connection() as conn:
        with conn.cursor() as cur:
            cycle_id = selected._ensure_cycle(cur, payload.site, payload.distribution_date)
            cur.execute(
                """
                select id,excel_evidence_uri,status,sent_at,generated_filename,drive_upload_status,drive_upload_error
                from accountant_submissions
                where upper(site)=upper(%s) and upper(accountant_code)=upper(%s)
                  and source_calculator_document_id=%s
                order by id desc limit 1
                """,
                (payload.site, accountant, payload.calculator_document_id),
            )
            existing = cur.fetchone()
            if existing and existing.get("excel_evidence_uri"):
                return {
                    **base,
                    "committed": True,
                    "duplicate": True,
                    "submissionId": existing["id"],
                    "driveUri": existing["excel_evidence_uri"],
                    "status": existing["status"],
                    "sentAt": existing["sent_at"],
                    "filename": existing.get("generated_filename") or artifact["filename"],
                    "driveUploadStatus": existing.get("drive_upload_status") or "UPLOADED",
                    "driveUploadError": existing.get("drive_upload_error"),
                }

            try:
                upload = upload_accountant_artifact(
                    kind="excel",
                    filename=artifact["filename"],
                    data=artifact["xlsx"],
                    mime_type=excel.XLSX_MIME,
                )
            except AccountantDriveUploadError as exc:
                error_text = str(exc)[:1500]
                cur.execute(
                    """
                    insert into accountant_submissions(
                      production_cycle_id,site,accountant_code,excel_evidence_uri,sent_at,status,
                      source_planning_snapshot_id,generated_filename,source_calculator_document_id,
                      source_plan_name,source_distribution_date,drive_upload_status,drive_upload_error,updated_at
                    ) values (%s,%s,%s,null,null,'EXCEL_READY_UPLOAD_FAILED',null,%s,%s,%s,%s,'FAILED',%s,now())
                    on conflict (site,accountant_code,source_calculator_document_id)
                    where source_calculator_document_id is not null
                    do update set generated_filename=excluded.generated_filename,
                      source_plan_name=excluded.source_plan_name,
                      source_distribution_date=excluded.source_distribution_date,
                      drive_upload_status='FAILED',drive_upload_error=excluded.drive_upload_error,
                      status=case when accountant_submissions.status='SENT' then accountant_submissions.status else 'EXCEL_READY_UPLOAD_FAILED' end,
                      updated_at=now()
                    returning id,status
                    """,
                    (
                        cycle_id,
                        payload.site,
                        accountant,
                        artifact["filename"],
                        payload.calculator_document_id,
                        artifact["planName"],
                        payload.distribution_date,
                        error_text,
                    ),
                )
                failed = cur.fetchone()
                conn.commit()
                return {
                    **base,
                    "committed": False,
                    "submissionId": failed["id"],
                    "status": failed["status"],
                    "driveUploadStatus": "FAILED",
                    "driveUploadError": error_text,
                    "retryable": True,
                }

            cur.execute(
                """
                insert into accountant_submissions(
                  production_cycle_id,site,accountant_code,excel_evidence_uri,sent_at,status,
                  source_planning_snapshot_id,generated_filename,source_calculator_document_id,
                  source_plan_name,source_distribution_date,drive_upload_status,drive_upload_error,updated_at
                ) values (%s,%s,%s,%s,null,'READY',null,%s,%s,%s,%s,'UPLOADED',null,now())
                on conflict (site,accountant_code,source_calculator_document_id)
                where source_calculator_document_id is not null
                do update set production_cycle_id=excluded.production_cycle_id,
                  excel_evidence_uri=excluded.excel_evidence_uri,generated_filename=excluded.generated_filename,
                  source_plan_name=excluded.source_plan_name,source_distribution_date=excluded.source_distribution_date,
                  drive_upload_status='UPLOADED',drive_upload_error=null,
                  status=case when accountant_submissions.status='SENT' then accountant_submissions.status else 'READY' end,
                  updated_at=now()
                returning id,status,sent_at
                """,
                (
                    cycle_id,
                    payload.site,
                    accountant,
                    upload["driveUri"],
                    artifact["filename"],
                    payload.calculator_document_id,
                    artifact["planName"],
                    payload.distribution_date,
                ),
            )
            saved = cur.fetchone()
            conn.commit()
            return {
                **base,
                "committed": True,
                "submissionId": saved["id"],
                "driveUri": upload["driveUri"],
                "driveFolderId": upload["folderId"],
                "usedFallbackDriveFolder": upload["usedFallbackFolder"],
                "status": saved["status"],
                "sentAt": saved["sent_at"],
                "driveUploadStatus": "UPLOADED",
            }


def named_download_selected_plan_excel(
    site: Literal["MAJA", "CEMPLANG"],
    distribution_date: date = Query(alias="distributionDate"),
    calculator_document_id: str = Query(alias="calculatorDocumentId", min_length=1),
    custom_filename: str | None = Query(default=None, alias="customFilename", max_length=160),
):
    selected.require_db()
    payload = NamedSelectedPlanExcelIn(
        site=site,
        distribution_date=distribution_date,
        calculator_document_id=calculator_document_id,
        custom_filename=custom_filename,
        commit=False,
    )
    artifact = _artifact(payload)
    return StreamingResponse(
        io.BytesIO(artifact["xlsx"]),
        media_type=excel.XLSX_MIME,
        headers={
            "Content-Disposition": f'attachment; filename="{artifact["filename"]}"',
            "Cache-Control": "no-store",
            "X-SPPG-Excel-Source": "calculator-daily-plan-single",
        },
    )


def install() -> None:
    global _INSTALLED
    if _INSTALLED:
        return

    # All accountant workbook creation paths use the polished renderer.
    excel._workbook_bytes = polished_workbook_bytes

    # Replace only the selected-plan Excel routes; invoice and Maker routes remain untouched.
    selected.router.routes[:] = [
        route for route in selected.router.routes
        if not (
            getattr(route, "path", "") == "/accountant-excel/from-selected-plan"
            and "POST" in (getattr(route, "methods", set()) or set())
        )
        and not (
            getattr(route, "path", "") == "/accountant-excel/download-selected-plan"
            and "GET" in (getattr(route, "methods", set()) or set())
        )
    ]
    selected.router.add_api_route(
        "/accountant-excel/from-selected-plan",
        named_selected_plan_excel,
        methods=["POST"],
        summary="Generate accountant Excel from one selected calculator plan",
    )
    selected.router.add_api_route(
        "/accountant-excel/download-selected-plan",
        named_download_selected_plan_excel,
        methods=["GET"],
        include_in_schema=False,
        summary="Download selected-plan accountant Excel",
    )
    _INSTALLED = True
