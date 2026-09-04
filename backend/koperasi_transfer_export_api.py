from __future__ import annotations

from datetime import date
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from backend.db import connection, database_ready

router = APIRouter(prefix="/v1", tags=["inventory-export"])


def _require_db() -> None:
    if not database_ready():
        raise HTTPException(503, "database unavailable")


def _destination(value: str) -> str:
    target = str(value or "").upper().strip()
    if target and target not in {"MAJA", "CEMPLANG"}:
        raise HTTPException(400, "destination must be MAJA or CEMPLANG")
    return target


@router.get("/inventory/koperasi-transfers/export.xlsx")
def export_koperasi_transfers_xlsx(
    from_date: date = Query(alias="fromDate"),
    to_date: date = Query(alias="toDate"),
    destination: str = "",
):
    """Export real XLSX for Gudang Koperasi deliveries, filtered by date and site."""
    _require_db()
    if to_date < from_date:
        raise HTTPException(400, "toDate must be on or after fromDate")
    target = _destination(destination)

    sql = """
        select date(coalesce(im.occurred_at,im.created_at)) transfer_date,
               upper(im.to_location) destination,
               im.item_name,im.qty,im.unit,
               po.po_code,gr.receipt_code,gr.reporter,
               im.id movement_id
        from inventory_movements im
        left join goods_receipts gr
          on im.source_type='GOODS_RECEIPT' and im.source_ref=('receipt:' || gr.id::text)
        left join purchase_orders po on po.id=gr.purchase_order_id
        where upper(coalesce(im.from_location,''))='KOPERASI'
          and upper(coalesce(im.to_location,'')) in ('MAJA','CEMPLANG')
          and upper(coalesce(im.movement_type,'')) in ('KOPERASI_STOCK_TRANSFER','PURCHASE_RECEIPT')
          and date(coalesce(im.occurred_at,im.created_at)) between %s and %s
    """
    params = [from_date, to_date]
    if target:
        sql += " and upper(im.to_location)=%s"
        params.append(target)
    sql += " order by transfer_date,upper(im.to_location),lower(im.item_name),im.id"

    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Kiriman Koperasi"
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "KIRIMAN BARANG GUDANG KOPERASI"
    title.font = Font(size=14, bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:H2")
    subtitle = ws["A2"]
    site_label = target or "MAJA + CEMPLANG"
    subtitle.value = f"Periode {from_date.strftime('%d-%m-%Y')} s.d. {to_date.strftime('%d-%m-%Y')} | Tujuan: {site_label}"
    subtitle.font = Font(italic=True)
    subtitle.alignment = Alignment(horizontal="center")

    headers = ["Tanggal Kirim", "Site Tujuan", "Barang", "Qty", "Satuan", "PO", "Bukti Penerimaan", "Penerima"]
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B8C2CC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column, label in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=column, value=label)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, start=5):
        values = [
            row["transfer_date"],
            row["destination"],
            row["item_name"],
            float(row["qty"] or 0),
            row["unit"] or "",
            row["po_code"] or "-",
            row["receipt_code"] or "-",
            row["reporter"] or "-",
        ]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=column, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=column in {3, 6, 7, 8})
        ws.cell(row=row_index, column=1).number_format = "dd-mmm-yyyy"
        ws.cell(row=row_index, column=4).number_format = "0.####"

    last_row = max(4, 4 + len(rows))
    ws.auto_filter.ref = f"A4:H{last_row}"
    widths = {"A": 15, "B": 15, "C": 34, "D": 13, "E": 12, "F": 25, "G": 22, "H": 22}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    if not rows:
        ws.merge_cells("A5:H5")
        empty = ws["A5"]
        empty.value = "Tidak ada kiriman pada filter yang dipilih."
        empty.alignment = Alignment(horizontal="center")
        empty.font = Font(italic=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    suffix = target or "SEMUA_SITE"
    filename = f"Kiriman_Koperasi_{from_date.isoformat()}_{to_date.isoformat()}_{suffix}.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )
