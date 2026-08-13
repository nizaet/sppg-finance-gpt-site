from __future__ import annotations

import io
import json
import os
from datetime import date
from typing import Any, Literal

from fastapi import APIRouter, HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from backend.db import connection, database_ready
from backend.google_services import GoogleServicesNotConfigured, upload_bytes_to_drive

router = APIRouter(tags=["accountant-excel"])

ACCOUNTANTS = {"MAJA": "TIARA", "CEMPLANG": "UYA"}
VENDOR_LABELS = {
    "HOLIL": "Haji Holil",
    "WIKIAN": "Wikian",
    "RUMAH_DUTA_PANGAN": "Rumah Duta Pangan",
    "HERU": "Heru",
    "DEDE": "Dede",
    "HAJI_BADRI": "Haji Badri",
    "KOPERASI": "Koperasi / Mungki",
}
BGN_RATE_KECIL = 8000
BGN_RATE_BESAR = 10000
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class AccountantExcelFromPlanningIn(BaseModel):
    site: Literal["MAJA", "CEMPLANG"]
    distribution_date: date
    planning_snapshot_id: int | None = None
    commit: bool = False


def require_db() -> None:
    if not database_ready():
        raise HTTPException(503, "database unavailable")


def _payload_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {}
    return {}


def _supplier_label(item: dict[str, Any]) -> str:
    source = _payload_dict(item.get("source_payload"))
    for key in ("supplier_title", "supplierTitle", "kategori_supplier", "category_supplier", "supplier"):
        value = str(source.get(key) or "").strip()
        if value:
            return value
    preferred = str(item.get("preferred_vendor_code") or "").upper().strip()
    if preferred:
        return VENDOR_LABELS.get(preferred, preferred)
    return str(item.get("category_code") or "").strip()


def _number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _extract_portions(payload: dict[str, Any]) -> tuple[float | None, float | None]:
    aliases_small = ("porsiKecil", "porsi_kecil", "smallPortions", "small_portions")
    aliases_large = ("porsiBesar", "porsi_besar", "largePortions", "large_portions")

    def first(keys):
        for key in keys:
            if key in payload and payload.get(key) not in (None, ""):
                return _number(payload.get(key))
        return None

    return first(aliases_small), first(aliases_large)


def _load_snapshot(cur, payload: AccountantExcelFromPlanningIn) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    if payload.planning_snapshot_id is not None:
        cur.execute(
            """select * from planning_snapshots
               where id=%s and upper(site)=upper(%s) and distribution_date=%s""",
            (payload.planning_snapshot_id, payload.site, payload.distribution_date),
        )
    else:
        cur.execute(
            """select * from planning_snapshots
               where upper(site)=upper(%s) and distribution_date=%s and status='ACTIVE'
               order by created_at desc limit 1""",
            (payload.site, payload.distribution_date),
        )
    snapshot = cur.fetchone()
    if not snapshot:
        raise HTTPException(404, "planning snapshot Kalkulator tidak ditemukan untuk site/tanggal ini")
    cur.execute(
        """select * from planning_snapshot_items
           where planning_snapshot_id=%s order by id""",
        (snapshot["id"],),
    )
    items = cur.fetchall()
    if not items:
        raise HTTPException(409, "planning snapshot tidak memiliki item belanja")
    return snapshot, items


def _build_rows(snapshot: dict[str, Any], items: list[dict[str, Any]]) -> tuple[list[list[Any]], float, float | None, float | None]:
    header = [
        "Item",
        "Jumlah",
        "Satuan",
        "Harga Satuan (Estimasi)",
        "Total Harga (Estimasi)",
        "Catatan",
        "Kategori / Supplier",
    ]
    normalized: list[dict[str, Any]] = []
    for item in items:
        q = _number(item.get("planned_qty"))
        price = _number(item.get("planning_price"))
        normalized.append({
            "item": str(item.get("item_name") or "").strip(),
            "qty": q,
            "unit": str(item.get("unit") or "").strip(),
            "price": price,
            "total": round(q * price, 2),
            "notes": str(item.get("notes") or "").strip(),
            "supplier": _supplier_label(item),
        })
    normalized.sort(key=lambda x: (x["supplier"].casefold(), x["item"].casefold()))

    rows: list[list[Any]] = [header]
    rows.extend([[x["item"], x["qty"], x["unit"], x["price"], x["total"], x["notes"], x["supplier"]] for x in normalized])
    grand_total = round(sum(x["total"] for x in normalized), 2)
    rows.append([])
    rows.append(["GRAND TOTAL", "", "", "", grand_total, "", ""])

    raw_payload = _payload_dict(snapshot.get("payload"))
    small, large = _extract_portions(raw_payload)
    pagu: float | None = None
    delta: float | None = None
    if small is not None or large is not None:
        pagu = round((small or 0) * BGN_RATE_KECIL + (large or 0) * BGN_RATE_BESAR, 2)
        delta = round(pagu - grand_total, 2)
        rows.append(["PAGU BGN", "", "", "", pagu, "", ""])
        rows.append(["SELISIH PAGU - ESTIMASI", "", "", "", delta, "", ""])

    return rows, grand_total, pagu, delta


def _workbook_bytes(rows: list[list[Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Belanja"

    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(color="FFFFFF", bold=True)
    summary_fill = PatternFill("solid", fgColor="DBEAFE")
    summary_font = Font(bold=True, color="1E3A8A")
    thin_gray = Side(style="thin", color="CBD5E1")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
        if not any(cell.value not in (None, "") for cell in row):
            continue
        is_summary = str(row[0].value or "").upper() in {"GRAND TOTAL", "PAGU BGN", "SELISIH PAGU - ESTIMASI"}
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_summary:
                cell.fill = summary_fill
                cell.font = summary_font

    # Keep values numeric in Excel while displaying Indonesian-style grouped numbers.
    for row_index in range(2, ws.max_row + 1):
        ws.cell(row=row_index, column=2).number_format = '#,##0.####'
        ws.cell(row=row_index, column=4).number_format = '#,##0'
        ws.cell(row=row_index, column=5).number_format = '#,##0'

    widths = [32, 14, 13, 23, 24, 34, 24]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"

    # Autofilter only covers the item table, stopping before blank/summary rows.
    last_data_row = 1
    for idx in range(2, ws.max_row + 1):
        if ws.cell(idx, 1).value in (None, "", "GRAND TOTAL"):
            break
        last_data_row = idx
    if last_data_row >= 2:
        ws.auto_filter.ref = f"A1:G{last_data_row}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@router.post("/accountant-excel/from-planning")
def accountant_excel_from_planning(payload: AccountantExcelFromPlanningIn) -> dict[str, Any]:
    require_db()
    accountant = ACCOUNTANTS[payload.site]

    with connection() as conn:
        with conn.cursor() as cur:
            snapshot, items = _load_snapshot(cur, payload)
            rows, grand_total, pagu, delta = _build_rows(snapshot, items)
            filename = f"daftar_belanja_{payload.distribution_date.isoformat()}_{payload.site.lower()}.xlsx"

            result = {
                "committed": False,
                "duplicate": False,
                "site": payload.site,
                "accountantCode": accountant,
                "planningSnapshotId": snapshot["id"],
                "productionCycleId": snapshot.get("production_cycle_id"),
                "distributionDate": payload.distribution_date.isoformat(),
                "filename": filename,
                "sheetName": "Belanja",
                "columns": rows[0],
                "itemCount": len(items),
                "grandTotal": grand_total,
                "paguBgn": pagu,
                "paguMinusEstimate": delta,
                "driveUri": None,
            }
            if not payload.commit:
                return result

            cur.execute(
                """select id,excel_evidence_uri,status,sent_at,generated_filename
                   from accountant_submissions
                   where upper(site)=upper(%s) and upper(accountant_code)=upper(%s)
                     and source_planning_snapshot_id=%s""",
                (payload.site, accountant, snapshot["id"]),
            )
            existing = cur.fetchone()
            if existing:
                result.update({
                    "committed": True,
                    "duplicate": True,
                    "submissionId": existing["id"],
                    "driveUri": existing["excel_evidence_uri"],
                    "status": existing["status"],
                    "sentAt": existing["sent_at"],
                    "filename": existing["generated_filename"] or filename,
                })
                return result

            xlsx = _workbook_bytes(rows)
            folder_id = os.getenv("SPPG_DRIVE_ACCOUNTANT_FOLDER_ID", "").strip()
            try:
                drive_uri = upload_bytes_to_drive(folder_id, filename, xlsx, XLSX_MIME)
            except GoogleServicesNotConfigured as exc:
                raise HTTPException(503, str(exc)) from exc

            cur.execute(
                """insert into accountant_submissions(
                     production_cycle_id,site,accountant_code,excel_evidence_uri,sent_at,status,
                     source_planning_snapshot_id,generated_filename
                   ) values (%s,%s,%s,%s,null,'READY',%s,%s)
                   returning id""",
                (
                    snapshot.get("production_cycle_id"), payload.site, accountant, drive_uri,
                    snapshot["id"], filename,
                ),
            )
            submission_id = cur.fetchone()["id"]
            conn.commit()
            result.update({
                "committed": True,
                "duplicate": False,
                "submissionId": submission_id,
                "driveUri": drive_uri,
                "status": "READY",
                "sentAt": None,
            })
            return result
